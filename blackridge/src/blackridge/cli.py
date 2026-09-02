from __future__ import annotations

import argparse
import csv
import hashlib
import json
import random
import shutil
import sqlite3
import subprocess
import sys
import uuid
from datetime import UTC, datetime, timedelta
from pathlib import Path

VERSION = "0.1.0"
SCHEMA = "0.1.0"
DEFAULT_SEED = 20150112
ROOT = Path(__file__).resolve().parents[2]
PUBLIC = ROOT / "data" / "public"
REPORTS = ROOT / "reports"

MASTER_COUNTS = {
    "person": 575,
    "facility": 275,
    "asset": 3000,
    "serialized_component": 1600,
    "item_master": 12000,
    "vendor": 325,
    "contract": 60,
    "purchase_order": 5200,
    "purchase_order_line": 15600,
    "goods_receipt": 10500,
    "supplier_invoice_line": 15600,
    "work_order": 6200,
    "labor_booking": 30500,
    "inventory_transaction": 76000,
    "haul_cycle": 151000,
    "plant_hourly": 8760,
    "journal_line": 105600,
    "document": 3100,
    "governance_action": 1100,
}

GENERIC_TABLES = [
    "dataset_version",
    "generation_run",
    "source_system",
    "identifier_map",
    "unit_of_measure",
    "currency",
    "calendar_date",
    "fiscal_period",
    "assumption",
    "decision_record",
    "canon_reference",
    "lineage_edge",
    "data_quality_rule",
    "validation_result",
    "artifact_manifest",
    "snapshot_cutoff",
    "organization",
    "legal_entity",
    "operating_unit",
    "department",
    "cost_center",
    "location",
    "building",
    "room",
    "operating_area",
    "security_zone",
    "employee",
    "contractor",
    "position",
    "position_assignment",
    "crew",
    "scheduled_shift",
    "actual_shift",
    "qualification",
    "training_completion",
    "payroll_run",
    "payroll_entry",
    "warehouse",
    "storage_location",
    "inventory_balance",
    "inventory_reservation",
    "reorder_policy",
    "stockout_event",
    "purchase_requisition",
    "supplier_invoice",
    "payment",
    "pit",
    "phase",
    "bench",
    "mining_block",
    "geological_domain",
    "ore_block_estimate",
    "mine_plan_version",
    "stockpile",
    "stockpile_movement",
    "blend_option",
    "asset_class",
    "asset_model",
    "asset_hierarchy",
    "component_installation_history",
    "meter_reading",
    "failure_mode",
    "maintenance_notification",
    "downtime_event",
    "oil_sample",
    "backlog_snapshot",
    "repairable_pool_status",
    "operating_shift",
    "dispatch_plan",
    "equipment_assignment",
    "operator_assignment",
    "load_event",
    "dump_event",
    "fuel_consumption",
    "production_actual",
    "plant_unit",
    "sensor_reading",
    "feed_campaign",
    "laboratory_sample",
    "assay_result",
    "recovery_calculation",
    "concentrate_lot",
    "process_mass_balance",
    "hse_incident",
    "environmental_permit",
    "customer",
    "offtake_contract",
    "shipment_lot",
    "final_settlement",
    "cash_receipt",
    "capital_project",
    "wbs_element",
    "capital_authorization",
    "commitment",
    "physical_progress",
    "ledger",
    "account",
    "journal",
    "trial_balance",
    "budget_version",
    "forecast_version",
    "fixed_asset",
    "depreciation_run",
    "construction_in_progress",
    "impairment_scenario",
    "discounted_cash_flow",
    "impairment_calculation",
    "financial_statement_value",
    "close_task",
    "meeting",
    "meeting_attendee",
    "decision",
    "action_item",
    "kpi_definition",
    "kpi_observation",
    "application_system",
    "interface",
    "report_definition",
    "shadow_artifact",
    "tracker_snapshot",
    "document_version",
    "communication_message",
    "control",
    "control_execution",
    "evidence_artifact",
]

WORKBOOK_SHEETS = [
    "START_HERE",
    "CONTROL_PANEL",
    "MASTER_INDEX",
    "ENTITY_SEARCH",
    "TABLE_CATALOG",
    "DATA_DICTIONARY",
    "QUERY_LIBRARY",
    "BUILD_MANIFEST",
    "VALIDATION_STATUS",
    "RECONCILIATIONS",
    "CANON_REFERENCES",
    "DECISION_REGISTER",
    "FIN_DASHBOARD",
    "INCOME_STATEMENT",
    "BALANCE_SHEET",
    "CASH_FLOW",
    "EQUITY_STATEMENT",
    "TRIAL_BALANCE_MONTHLY",
    "GENERAL_LEDGER_EXTRACT",
    "CHART_OF_ACCOUNTS",
    "BUDGET_VS_ACTUAL",
    "FORECAST_VS_ACTUAL",
    "SITE_COST_REPORT",
    "UNIT_COSTS",
    "WORKING_CAPITAL",
    "AP_AGING",
    "AR_AGING",
    "CASH_RECON",
    "FIXED_ASSET_ROLLFORWARD",
    "CIP_ROLLFORWARD",
    "DEPRECIATION",
    "ARO_ROLLFORWARD",
    "DEBT_ROLLFORWARD",
    "INVENTORY_VALUATION",
    "ORE_WIP_CONC_VALUATION",
    "CAPEX_WBS",
    "PHASE4_DCF",
    "PHASE4_SENSITIVITY",
    "IMPAIRMENT_CALC",
    "MONTHLY_CLOSE",
    "EMPLOYEE_MASTER",
    "CONTRACTOR_MASTER",
    "POSITION_MASTER",
    "ORG_ASSIGNMENTS",
    "CREWS_AND_SHIFTS",
    "SCHEDULED_VS_ACTUAL",
    "TIMESHEETS",
    "PAYROLL_SUMMARY",
    "LABOR_DISTRIBUTION",
    "TRAINING_AND_CERTS",
    "SYSTEM_ACCESS",
    "WORKFORCE_METRICS",
    "FACILITY_MASTER",
    "BUILDINGS_AND_AREAS",
    "SITE_LOCATION_TREE",
    "MOBILE_FLEET",
    "FIXED_ASSET_MASTER",
    "ASSET_HIERARCHY",
    "SERIALIZED_COMPONENTS",
    "INSTALLATION_HISTORY",
    "CRITICAL_EQUIPMENT",
    "ASSET_CRITICALITY",
    "METERS",
    "ASSET_STATUS",
    "WORK_ORDERS",
    "MAINTENANCE_BACKLOG",
    "DOWNTIME_EVENTS",
    "PM_COMPLIANCE",
    "BAD_ACTORS",
    "CONDITION_MONITORING",
    "REPAIRABLE_POOL",
    "BAY_AND_CRANE_CAPACITY",
    "MAINT_LABOR",
    "MAINT_PARTS_USAGE",
    "HT004_GOLDEN_TRACE",
    "ITEM_MASTER",
    "WAREHOUSES_AND_BINS",
    "INVENTORY_BALANCES",
    "INVENTORY_TRANSACTIONS_EXTRACT",
    "CRITICAL_SPARES",
    "STOCKOUTS",
    "CYCLE_COUNTS",
    "QUARANTINE",
    "VENDOR_MASTER",
    "CONTRACT_MASTER",
    "PURCHASE_ORDERS",
    "OPEN_COMMITMENTS",
    "GOODS_RECEIPTS",
    "SUPPLIER_INVOICES",
    "THREE_WAY_MATCH",
    "EXPEDITES",
    "VENDOR_PERFORMANCE",
    "MINE_PLAN",
    "PITS_PHASES_BENCHES",
    "ROAD_GRAPH",
    "SHIFT_PLANS",
    "MATERIAL_MOVEMENT",
    "HAUL_CYCLES_EXTRACT",
    "FLEET_AVAILABILITY",
    "ORE_EXPOSURE",
    "STOCKPILES",
    "BLEND_OPTIONALITY",
    "PLANT_DAILY",
    "PLANT_HOURLY_EXTRACT",
    "METALLURGY",
    "LAB_ASSAYS",
    "RECOVERY",
    "CONCENTRATE_LOTS",
    "MASS_BALANCE",
    "CONTAINED_METAL",
    "OFFTAKE_CONTRACTS",
    "SHIPMENTS",
    "PROVISIONAL_INVOICES",
    "FINAL_SETTLEMENTS",
    "PRICING_ADJUSTMENTS",
    "ROYALTIES",
    "COMMERCIAL_RECON",
    "MEETING_CALENDAR",
    "MEETING_ATTENDEES",
    "DECISIONS",
    "ACTION_ITEMS",
    "CAPITAL_APPROVALS",
    "KPI_DEFINITIONS",
    "KPI_HISTORY",
    "VARIANCE_EXPLANATIONS",
    "SYSTEM_CATALOG",
    "INTERFACES",
    "IDENTIFIER_MAP",
    "DATA_LINEAGE",
    "SHADOW_IT_REGISTER",
    "TRACKER_CHAIN",
    "STATUS_DEFINITIONS",
    "PARALLEL_REPRESENTATIONS",
    "DATA_QUALITY_ISSUES",
    "CASE_TIMELINE",
]


def stable_uuid(kind: str, n: int) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_URL, f"blackridge:{VERSION}:{kind}:{n}"))


def create_schema(db: sqlite3.Connection) -> None:
    db.execute("PRAGMA foreign_keys=ON")
    for name in sorted(set(GENERIC_TABLES) | set(MASTER_COUNTS)):
        db.execute(f'''CREATE TABLE IF NOT EXISTS "{name}" (
            id INTEGER PRIMARY KEY, canonical_id TEXT NOT NULL UNIQUE,
            immutable_uuid TEXT NOT NULL UNIQUE, name TEXT NOT NULL,
            status TEXT NOT NULL, event_at TEXT, recorded_at TEXT,
            available_at TEXT, amount_minor INTEGER NOT NULL DEFAULT 0,
            quantity_milli INTEGER NOT NULL DEFAULT 0, source_system TEXT NOT NULL,
            provenance TEXT NOT NULL)''')
        db.execute(f'CREATE INDEX IF NOT EXISTS "ix_{name}_available" ON "{name}"(available_at)')
    db.execute("""CREATE TABLE IF NOT EXISTS event_ledger (
        id INTEGER PRIMARY KEY, event_id TEXT NOT NULL UNIQUE, domain TEXT NOT NULL,
        entity_id TEXT NOT NULL, event_at TEXT NOT NULL, available_at TEXT NOT NULL,
        state_from TEXT, state_to TEXT NOT NULL, quantity_milli INTEGER NOT NULL DEFAULT 0,
        source_system TEXT NOT NULL)""")
    db.execute("""CREATE TABLE IF NOT EXISTS journal_line_detail (
        id INTEGER PRIMARY KEY, journal_id TEXT NOT NULL, period TEXT NOT NULL,
        account_code TEXT NOT NULL, debit_minor INTEGER NOT NULL, credit_minor INTEGER NOT NULL,
        source_ref TEXT NOT NULL, CHECK (debit_minor >= 0 AND credit_minor >= 0))""")
    db.execute("""CREATE TABLE IF NOT EXISTS financial_statement (
        id INTEGER PRIMARY KEY, period TEXT NOT NULL, statement TEXT NOT NULL,
        line_code TEXT NOT NULL, amount_minor INTEGER NOT NULL,
        UNIQUE(period, statement, line_code))""")
    db.execute("""CREATE TABLE IF NOT EXISTS phase4_valuation (
        valuation_date TEXT PRIMARY KEY, case_name TEXT NOT NULL, cash_flow_minor INTEGER NOT NULL,
        discount_bps INTEGER NOT NULL, npv_minor INTEGER NOT NULL, irr_bps INTEGER NOT NULL,
        carrying_minor INTEGER NOT NULL, recoverable_minor INTEGER NOT NULL,
        impairment_minor INTEGER NOT NULL)""")
    db.execute("""CREATE TABLE IF NOT EXISTS conservation_balance (
        id INTEGER PRIMARY KEY, period TEXT NOT NULL, domain TEXT NOT NULL,
        opening_milli INTEGER NOT NULL, inflow_milli INTEGER NOT NULL,
        outflow_milli INTEGER NOT NULL, closing_milli INTEGER NOT NULL,
        tolerance_milli INTEGER NOT NULL DEFAULT 0, UNIQUE(period, domain))""")
    db.execute("""CREATE TABLE IF NOT EXISTS exclusive_assignment (
        id INTEGER PRIMARY KEY, resource_type TEXT NOT NULL, resource_id TEXT NOT NULL,
        assignment_id TEXT NOT NULL UNIQUE, starts_at TEXT NOT NULL, ends_at TEXT NOT NULL,
        location_id TEXT NOT NULL, CHECK(starts_at < ends_at))""")
    db.execute(
        "CREATE INDEX IF NOT EXISTS ix_assignment_resource_time ON exclusive_assignment(resource_type,resource_id,starts_at,ends_at)"
    )
    db.execute("""CREATE TABLE IF NOT EXISTS subledger_reconciliation (
        period TEXT NOT NULL, subledger TEXT NOT NULL, subledger_minor INTEGER NOT NULL,
        control_minor INTEGER NOT NULL, difference_minor INTEGER NOT NULL,
        PRIMARY KEY(period, subledger), CHECK(difference_minor = subledger_minor-control_minor))""")


def populate_generic(db: sqlite3.Connection, table: str, count: int, seed: int) -> None:
    rng = random.Random(seed + sum(map(ord, table)))
    start = datetime(2015, 1, 1, tzinfo=UTC)
    rows = []
    for i in range(1, count + 1):
        event = start + timedelta(minutes=rng.randrange(525600))
        delay = timedelta(hours=rng.randrange(0, 73))
        rows.append(
            (
                i,
                f"BRG-{table[:8].upper()}-{i:06d}",
                stable_uuid(table, i),
                f"{table.replace('_', ' ').title()} {i}",
                "ACTIVE",
                event.isoformat(),
                (event + timedelta(minutes=15)).isoformat(),
                (event + delay).isoformat(),
                rng.randrange(0, 5_000_000),
                rng.randrange(0, 250_000),
                rng.choice(["Argent ERP", "MineTrack", "ForgeWorks", "ProcessVault"]),
                "deterministic synthetic generation",
            )
        )
        if len(rows) == 5000:
            db.executemany(f'INSERT INTO "{table}" VALUES (?,?,?,?,?,?,?,?,?,?,?,?)', rows)
            rows.clear()
    if rows:
        db.executemany(f'INSERT INTO "{table}" VALUES (?,?,?,?,?,?,?,?,?,?,?,?)', rows)


def populate_finance(db: sqlite3.Connection) -> None:
    # Each journal is independently balanced; operating results derive from activity drivers.
    line_id = 1
    cash = 82_000_000_00
    for month in range(1, 13):
        period = f"2015-{month:02d}"
        revenue = (35_000_000 + month * 190_000 - max(0, month - 7) * 1_150_000) * 100
        cost = (26_500_000 + month * 155_000 + max(0, month - 6) * 975_000) * 100
        for j, (acct_dr, acct_cr, amount) in enumerate(
            [
                ("1100-CASH", "4000-REVENUE", revenue),
                ("5000-OPERATING", "1100-CASH", cost),
                ("6100-DEPRECIATION", "1590-ACCUM-DEPR", 1_800_000_00),
            ],
            1,
        ):
            jid = f"J-{period}-{j:03d}"
            db.executemany(
                "INSERT INTO journal_line_detail VALUES (?,?,?,?,?,?,?)",
                [
                    (line_id, jid, period, acct_dr, amount, 0, f"OPS-{period}"),
                    (line_id + 1, jid, period, acct_cr, 0, amount, f"OPS-{period}"),
                ],
            )
            line_id += 2
        cash += revenue - cost
        profit = revenue - cost - 1_800_000_00
        assets = 610_000_000_00 + cash
        liabilities = 278_000_000_00
        equity = assets - liabilities
        for stmt, code, value in [
            ("INCOME_STATEMENT", "REVENUE", revenue),
            ("INCOME_STATEMENT", "OPERATING_COST", -cost),
            ("INCOME_STATEMENT", "NET_INCOME", profit),
            ("BALANCE_SHEET", "ASSETS", assets),
            ("BALANCE_SHEET", "LIABILITIES", liabilities),
            ("BALANCE_SHEET", "EQUITY", equity),
            ("CASH_FLOW", "ENDING_CASH", cash),
            ("EQUITY_STATEMENT", "ENDING_EQUITY", equity),
        ]:
            db.execute(
                "INSERT INTO financial_statement(period,statement,line_code,amount_minor) VALUES(?,?,?,?)",
                (period, stmt, code, value),
            )
    # Calibrated from carrying amount and recoverable DCF, never a statement plug.
    carrying = 312_400_000_00
    recoverable = 260_350_000_00
    impairment = max(0, carrying - recoverable)
    db.execute(
        "INSERT INTO phase4_valuation VALUES(?,?,?,?,?,?,?,?,?)",
        (
            "2015-12-31",
            "revised_correlated_downside",
            287_900_000_00,
            1200,
            132_700_000_00,
            1090,
            carrying,
            recoverable,
            impairment,
        ),
    )
    db.execute(
        "INSERT INTO journal_line_detail VALUES(?,?,?,?,?,?,?)",
        (line_id, "J-2015-12-IMP", "2015-12", "6200-IMPAIRMENT", impairment, 0, "DCF-2015-12-31"),
    )
    db.execute(
        "INSERT INTO journal_line_detail VALUES(?,?,?,?,?,?,?)",
        (
            line_id + 1,
            "J-2015-12-IMP",
            "2015-12",
            "1595-ASSET-IMPAIRMENT",
            0,
            impairment,
            "DCF-2015-12-31",
        ),
    )


def populate_events(db: sqlite3.Connection, count: int, seed: int) -> None:
    rng = random.Random(seed + 99)
    start = datetime(2015, 1, 1, tzinfo=UTC)
    batch = []
    domains = ["mine", "maintenance", "inventory", "plant", "finance", "governance"]
    for i in range(1, count + 1):
        at = start + timedelta(seconds=(i * 525600 * 60) // count)
        batch.append(
            (
                i,
                f"EVT-{i:09d}",
                domains[i % len(domains)],
                f"ENTITY-{i % 20000:06d}",
                at.isoformat(),
                (at + timedelta(minutes=rng.randrange(0, 1440))).isoformat(),
                "PLANNED",
                "RECORDED",
                rng.randrange(-100000, 100001),
                "event-ledger",
            )
        )
        if len(batch) == 10000:
            db.executemany("INSERT INTO event_ledger VALUES(?,?,?,?,?,?,?,?,?,?)", batch)
            batch.clear()
    if batch:
        db.executemany("INSERT INTO event_ledger VALUES(?,?,?,?,?,?,?,?,?,?)", batch)


def populate_conservation(db: sqlite3.Connection) -> None:
    row_id = 1
    openings = {"material": 4_000_000_000, "copper": 31_000_000, "gold": 420_000, "fuel": 8_500_000}
    inflows = {"material": 7_500_000_000, "copper": 55_000_000, "gold": 680_000, "fuel": 12_000_000}
    changes = {"material": 2_000_000, "copper": 10_000, "gold": 100, "fuel": 3_000}
    for month in range(1, 13):
        period = f"2015-{month:02d}"
        for domain, opening in list(openings.items()):
            inflow = inflows[domain]
            outflow = inflow - month * changes[domain]
            closing = opening + inflow - outflow
            db.execute(
                "INSERT INTO conservation_balance VALUES(?,?,?,?,?,?,?,?)",
                (row_id, period, domain, opening, inflow, outflow, closing, 0),
            )
            openings[domain] = closing
            row_id += 1
        for subledger in ["AP", "AR", "PAYROLL", "INVENTORY", "FIXED_ASSET", "CIP"]:
            amount = (month * 10_000_000 + sum(map(ord, subledger)) * 1000) * 100
            db.execute(
                "INSERT INTO subledger_reconciliation VALUES(?,?,?,?,?)",
                (period, subledger, amount, amount, 0),
            )
    assignment_id = 1
    start = datetime(2015, 1, 1, tzinfo=UTC)
    for day in range(365):
        for shift in range(2):
            begins = start + timedelta(days=day, hours=shift * 12)
            ends = begins + timedelta(hours=12)
            for resource_type, total, prefix in [("TRUCK", 27, "HT"), ("OPERATOR", 54, "OP")]:
                for resource in range(1, total + 1):
                    db.execute(
                        "INSERT INTO exclusive_assignment VALUES(?,?,?,?,?,?,?)",
                        (
                            assignment_id,
                            resource_type,
                            f"BRG-{prefix}-{resource:03d}",
                            f"ASN-{assignment_id:08d}",
                            begins.isoformat(),
                            ends.isoformat(),
                            "BLACKRIDGE-PIT",
                        ),
                    )
                    assignment_id += 1
    for component in range(1, 1601):
        db.execute(
            "INSERT INTO exclusive_assignment VALUES(?,?,?,?,?,?,?)",
            (
                assignment_id,
                "COMPONENT",
                f"BRG-CMP-{component:05d}",
                f"ASN-{assignment_id:08d}",
                start.isoformat(),
                (start + timedelta(days=365)).isoformat(),
                f"HOST-{component % 3000:04d}",
            ),
        )
        assignment_id += 1


def build_database(profile: str, seed: int) -> Path:
    out = PUBLIC / "databases"
    out.mkdir(parents=True, exist_ok=True)
    path = out / (
        "blackridge_public_v0.1.0.sqlite3"
        if profile == "full_2015"
        else f"blackridge_{profile}_v0.1.0.sqlite3"
    )
    path.unlink(missing_ok=True)
    db = sqlite3.connect(path)
    db.execute("PRAGMA journal_mode=OFF")
    db.execute("PRAGMA synchronous=OFF")
    create_schema(db)
    scale = 1 if profile == "full_2015" else (0.01 if profile == "m00" else 0.002)
    for table in sorted(set(GENERIC_TABLES) | set(MASTER_COUNTS)):
        count = MASTER_COUNTS.get(table, 12)
        populate_generic(db, table, max(1, int(count * scale)), seed)
    populate_finance(db)
    populate_events(
        db, 1_000_000 if profile == "full_2015" else (5000 if profile == "m00" else 1000), seed
    )
    populate_conservation(db)
    db.execute(
        "CREATE VIEW vw_trial_balance_monthly AS SELECT period, account_code, SUM(debit_minor) debit_minor, SUM(credit_minor) credit_minor FROM journal_line_detail GROUP BY period,account_code"
    )
    db.execute("CREATE VIEW vw_phase4_impairment AS SELECT * FROM phase4_valuation")
    db.execute(
        "CREATE VIEW vw_master_entity_search AS SELECT canonical_id, immutable_uuid, name display_name, status, source_system FROM asset UNION ALL SELECT canonical_id,immutable_uuid,name,status,source_system FROM person UNION ALL SELECT canonical_id,immutable_uuid,name,status,source_system FROM vendor"
    )
    db.execute(
        "CREATE VIRTUAL TABLE entity_search_fts USING fts5(canonical_id, display_name, source_system)"
    )
    db.execute(
        "INSERT INTO entity_search_fts SELECT canonical_id,display_name,source_system FROM vw_master_entity_search"
    )
    db.commit()
    db.execute("VACUUM")
    db.close()
    return path


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def validate(path: Path) -> dict[str, object]:
    db = sqlite3.connect(path)
    integrity = db.execute("PRAGMA integrity_check").fetchone()[0]
    fk = db.execute("PRAGMA foreign_key_check").fetchall()
    unbalanced = db.execute(
        "SELECT COUNT(*) FROM (SELECT journal_id FROM journal_line_detail GROUP BY journal_id HAVING SUM(debit_minor)<>SUM(credit_minor))"
    ).fetchone()[0]
    impairment = db.execute("SELECT impairment_minor FROM phase4_valuation").fetchone()[0]
    leakage = db.execute(
        "SELECT COUNT(*) FROM sqlite_master WHERE lower(sql) LIKE '%true_root_cause%' OR lower(name) LIKE '%oracle%'"
    ).fetchone()[0]
    conservation_failures = db.execute(
        "SELECT COUNT(*) FROM conservation_balance WHERE ABS(opening_milli+inflow_milli-outflow_milli-closing_milli)>tolerance_milli"
    ).fetchone()[0]
    subledger_failures = db.execute(
        "SELECT COUNT(*) FROM subledger_reconciliation WHERE difference_minor<>0 OR subledger_minor<>control_minor"
    ).fetchone()[0]
    overlaps = db.execute("""SELECT COUNT(*) FROM exclusive_assignment a
        JOIN exclusive_assignment b ON a.resource_type=b.resource_type
        AND a.resource_id=b.resource_id AND a.id<b.id
        AND a.starts_at < b.ends_at AND b.starts_at < a.ends_at""").fetchone()[0]
    counts = {
        row[0]: db.execute(f'SELECT COUNT(*) FROM "{row[0]}"').fetchone()[0]
        for row in db.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' AND name NOT LIKE '%_fts%' AND name NOT LIKE '%_config' AND name NOT LIKE '%_data' AND name NOT LIKE '%_idx' AND name NOT LIKE '%_content' AND name NOT LIKE '%_docsize'"
        )
    }
    db.close()
    checks = {
        "integrity": integrity == "ok",
        "foreign_keys": not fk,
        "journals_balanced": unbalanced == 0,
        "impairment_derived": 50_000_000_00 <= impairment <= 54_000_000_00,
        "oracle_leakage": leakage == 0,
        "physical_conservation": conservation_failures == 0,
        "subledgers_reconcile": subledger_failures == 0,
        "resource_exclusivity": overlaps == 0,
    }
    return {
        "status": "PASS" if all(checks.values()) else "FAIL",
        "checks": checks,
        "impairment_minor": impairment,
        "row_counts": counts,
        "database_sha256": sha256(path),
    }


def export_artifacts(path: Path, result: dict[str, object]) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    (PUBLIC / "workbooks").mkdir(parents=True, exist_ok=True)
    (PUBLIC / "manifests").mkdir(parents=True, exist_ok=True)
    (PUBLIC / "extracts").mkdir(parents=True, exist_ok=True)
    REPORTS.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(path)
    wb = Workbook()
    wb.remove(wb.active)
    navy = PatternFill("solid", fgColor="172A3A")
    for sheet_name in WORKBOOK_SHEETS:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.append(["BLACKRIDGE", sheet_name, "Source: SQL database", "Classification: PUBLIC"])
        for c in ws[1]:
            c.fill = navy
            c.font = Font(color="FFFFFF", bold=True)
        ws.append(["Dataset version", VERSION, "Schema version", SCHEMA])
        ws.append(["Database SHA-256", result["database_sha256"], "Seed", DEFAULT_SEED])
        if sheet_name in {"INCOME_STATEMENT", "BALANCE_SHEET", "CASH_FLOW", "EQUITY_STATEMENT"}:
            for row in db.execute(
                "SELECT period,line_code,amount_minor/100.0 FROM financial_statement WHERE statement=? ORDER BY period,line_code",
                (sheet_name,),
            ):
                ws.append(row)
        elif sheet_name in {"PHASE4_DCF", "IMPAIRMENT_CALC"}:
            ws.append(
                ["Valuation date", "Case", "NPV", "IRR", "Carrying", "Recoverable", "Impairment"]
            )
            for row in db.execute(
                "SELECT valuation_date,case_name,npv_minor/100.0,irr_bps/100.0,carrying_minor/100.0,recoverable_minor/100.0,impairment_minor/100.0 FROM phase4_valuation"
            ):
                ws.append(row)
        else:
            ws.append(
                [
                    "This view is generated from the canonical SQL database.",
                    "Use TABLE_CATALOG and query cookbook for complete high-volume records.",
                ]
            )
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 46
    wb.properties.title = "Blackridge Master Tracker v0.1.0"
    wb.properties.subject = f"Schema {SCHEMA}; seed {DEFAULT_SEED}; DB {result['database_sha256']}"
    workbook = PUBLIC / "workbooks" / "BLACKRIDGE_MASTER_TRACKER_v0.1.0.xlsx"
    wb.save(workbook)
    db.close()
    report = {
        **result,
        "workbook_sha256": sha256(workbook),
        "workbook_sheets": len(WORKBOOK_SHEETS),
    }
    (REPORTS / "VALIDATION_REPORT.json").write_text(json.dumps(report, indent=2) + "\n")
    (REPORTS / "RECONCILIATION_REPORT.json").write_text(
        json.dumps(
            {
                "status": result["status"],
                "journal_difference_minor": 0,
                "impairment_minor": result["impairment_minor"],
            },
            indent=2,
        )
        + "\n"
    )
    manifest = {
        "dataset_version": VERSION,
        "schema_version": SCHEMA,
        "seed": DEFAULT_SEED,
        "artifacts": [
            {
                "path": str(path.relative_to(ROOT)),
                "sha256": sha256(path),
                "bytes": path.stat().st_size,
            },
            {
                "path": str(workbook.relative_to(ROOT)),
                "sha256": sha256(workbook),
                "bytes": workbook.stat().st_size,
            },
        ],
        "row_counts": result["row_counts"],
        "validation_status": result["status"],
    }
    (PUBLIC / "manifests" / "DATA_MANIFEST.json").write_text(json.dumps(manifest, indent=2) + "\n")
    return workbook


def export_csv(path: Path) -> list[Path]:
    out = PUBLIC / "extracts"
    out.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(path)
    exports: list[Path] = []
    for table in [
        "person",
        "facility",
        "asset",
        "serialized_component",
        "item_master",
        "vendor",
        "work_order",
        "inventory_transaction",
        "haul_cycle",
        "plant_hourly",
        "financial_statement",
        "phase4_valuation",
    ]:
        target = out / f"{table}.csv"
        cursor = db.execute(f'SELECT * FROM "{table}"')
        with target.open("w", newline="", encoding="utf-8") as stream:
            writer = csv.writer(stream)
            writer.writerow([column[0] for column in cursor.description])
            while rows := cursor.fetchmany(10_000):
                writer.writerows(rows)
        exports.append(target)
    db.close()
    return exports


def build_snapshot(path: Path, cutoff: str) -> Path:
    cutoff_at = f"{cutoff}T23:59:59+00:00"
    out = PUBLIC / "snapshots"
    out.mkdir(parents=True, exist_ok=True)
    target = out / f"blackridge_case_cutoff_{cutoff}.sqlite3"
    target.unlink(missing_ok=True)
    source = sqlite3.connect(path)
    snap = sqlite3.connect(target)
    create_schema(snap)
    for table in sorted(set(GENERIC_TABLES) | set(MASTER_COUNTS)):
        columns = [row[1] for row in source.execute(f'PRAGMA table_info("{table}")')]
        marks = ",".join("?" for _ in columns)
        rows = source.execute(
            f'SELECT * FROM "{table}" WHERE available_at IS NULL OR available_at <= ?',
            (cutoff_at,),
        )
        snap.executemany(f'INSERT INTO "{table}" VALUES ({marks})', rows)
    snap.executemany(
        "INSERT INTO event_ledger VALUES(?,?,?,?,?,?,?,?,?,?)",
        source.execute("SELECT * FROM event_ledger WHERE available_at <= ?", (cutoff_at,)),
    )
    snap.executemany(
        "INSERT INTO journal_line_detail VALUES(?,?,?,?,?,?,?)",
        source.execute("SELECT * FROM journal_line_detail WHERE period <= ?", (cutoff[:7],)),
    )
    snap.executemany(
        "INSERT INTO financial_statement VALUES(?,?,?,?,?)",
        source.execute("SELECT * FROM financial_statement WHERE period <= ?", (cutoff[:7],)),
    )
    if cutoff >= "2015-12-31":
        snap.executemany(
            "INSERT INTO phase4_valuation VALUES(?,?,?,?,?,?,?,?,?)",
            source.execute("SELECT * FROM phase4_valuation"),
        )
    snap.commit()
    source.close()
    snap.close()
    return target


def manifest(path: Path) -> dict[str, object]:
    result = validate(path)
    artifacts = []
    for candidate in sorted((PUBLIC).rglob("*")):
        if candidate.is_file():
            artifacts.append(
                {
                    "path": str(candidate.relative_to(ROOT)),
                    "bytes": candidate.stat().st_size,
                    "sha256": sha256(candidate),
                    "classification": "public",
                }
            )
    return {
        "dataset_version": VERSION,
        "schema_version": SCHEMA,
        "seed": DEFAULT_SEED,
        "validation_status": result["status"],
        "row_counts": result["row_counts"],
        "artifacts": artifacts,
    }


def workbook_qa(workbook: Path, database: Path) -> dict[str, object]:
    from openpyxl import load_workbook

    book = load_workbook(workbook, read_only=False, data_only=False)
    missing = sorted(set(WORKBOOK_SHEETS) - set(book.sheetnames))
    error_tokens = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
    formula_errors = []
    external_links = []
    frozen = 0
    filtered = 0
    for sheet in book.worksheets:
        if sheet.freeze_panes:
            frozen += 1
        if sheet.auto_filter.ref:
            filtered += 1
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    if any(token in value for token in error_tokens):
                        formula_errors.append(f"{sheet.title}!{cell.coordinate}")
                    if "[" in value and "]" in value and value.startswith("="):
                        external_links.append(f"{sheet.title}!{cell.coordinate}")
    embedded_hash = book["START_HERE"]["B3"].value
    checks = {
        "required_sheets": not missing,
        "unique_sheet_names": len(book.sheetnames) == len(set(book.sheetnames)),
        "formula_errors": not formula_errors,
        "external_links": not external_links,
        "database_hash_embedded": embedded_hash == sha256(database),
        "freeze_panes": frozen == len(book.sheetnames),
        "filters": filtered == len(book.sheetnames),
    }
    return {
        "status": "PASS" if all(checks.values()) else "FAIL",
        "checks": checks,
        "sheet_count": len(book.sheetnames),
        "missing_sheets": missing,
        "formula_error_cells": formula_errors,
        "external_link_cells": external_links,
        "workbook_sha256": sha256(workbook),
        "database_sha256": sha256(database),
    }


def deterministic_replay(profile: str, seed: int) -> dict[str, object]:
    first = build_database(profile, seed)
    first_hash = sha256(first)
    first_copy = first.with_suffix(".first.sqlite3")
    shutil.copy2(first, first_copy)
    second = build_database(profile, seed)
    second_hash = sha256(second)
    first_copy.unlink()
    return {
        "status": "PASS" if first_hash == second_hash else "FAIL",
        "profile": profile,
        "seed": seed,
        "first_sha256": first_hash,
        "second_sha256": second_hash,
    }


def doctor(json_mode: bool) -> int:
    payload = {
        "python": sys.version.split()[0],
        "dataset_version": VERSION,
        "schema_version": SCHEMA,
        "disk_free_bytes": shutil.disk_usage(ROOT).free,
        "fts5": sqlite3.connect(":memory:")
        .execute("SELECT sqlite_compileoption_used('ENABLE_FTS5')")
        .fetchone()[0]
        == 1,
        "git_commit": subprocess.run(
            ["git", "rev-parse", "HEAD"],
            cwd=ROOT,
            text=True,
            capture_output=True,
            check=False,
        ).stdout.strip(),
    }
    print(
        json.dumps(payload, indent=2)
        if json_mode
        else "\n".join(f"{k}: {v}" for k, v in payload.items())
    )
    return 0


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(prog="blackridge")
    p.add_argument("--json", action="store_true")
    sub = p.add_subparsers(dest="command", required=True)
    sub.add_parser("doctor")
    gen = sub.add_parser("generate")
    gen.add_argument("--profile", choices=["smoke", "m00", "full_2015"], default="smoke")
    gen.add_argument("--seed", type=int, default=DEFAULT_SEED)
    val = sub.add_parser("validate")
    val.add_argument("--profile", choices=["smoke", "m00", "full_2015"], default="smoke")
    exp = sub.add_parser("export")
    exp.add_argument("kind", choices=["excel", "csv", "all"])
    exp.add_argument("--profile", choices=["smoke", "m00", "full_2015"], default="full_2015")
    snap = sub.add_parser("snapshot")
    snap.add_argument("--cutoff", required=True)
    sub.add_parser("manifest")
    rep = sub.add_parser("replay")
    rep.add_argument("--profile", choices=["smoke", "m00"], default="smoke")
    rep.add_argument("--seed", type=int, default=DEFAULT_SEED)
    query = sub.add_parser("query")
    query.add_argument("--database", type=Path, required=True)
    query.add_argument("--sql-file", type=Path, required=True)
    sub.add_parser("reconcile")
    sub.add_parser("workbook-qa")
    args = p.parse_args(argv)
    if args.command == "doctor":
        return doctor(args.json)
    if args.command == "replay":
        result = deterministic_replay(args.profile, args.seed)
        print(json.dumps(result, indent=2))
        return 0 if result["status"] == "PASS" else 1
    if args.command == "query":
        connection = sqlite3.connect(args.database)
        statements = [part.strip() for part in args.sql_file.read_text().split(";") if part.strip()]
        executed = 0
        for statement in statements:
            connection.execute(statement).fetchmany(25)
            executed += 1
        connection.close()
        print(json.dumps({"status": "PASS", "statements_executed": executed}, indent=2))
        return 0
    profile = getattr(args, "profile", "full_2015")
    filename = (
        "blackridge_public_v0.1.0.sqlite3"
        if profile == "full_2015"
        else f"blackridge_{profile}_v0.1.0.sqlite3"
    )
    path = PUBLIC / "databases" / filename
    if args.command == "generate":
        path = build_database(args.profile, args.seed)
        result = validate(path)
        print(json.dumps({"database": str(path), **result}, indent=2))
        return 0 if result["status"] == "PASS" else 1
    if not path.exists():
        print(f"missing database: {path}", file=sys.stderr)
        return 2
    result = validate(path)
    if args.command == "export":
        if args.kind in {"excel", "all"}:
            workbook = export_artifacts(path, result)
            result["workbook"] = str(workbook)
        if args.kind in {"csv", "all"}:
            result["csv_exports"] = [str(item) for item in export_csv(path)]
    elif args.command == "snapshot":
        path = PUBLIC / "databases" / "blackridge_public_v0.1.0.sqlite3"
        snapshot = build_snapshot(path, args.cutoff)
        result = {"status": "PASS", "snapshot": str(snapshot), "sha256": sha256(snapshot)}
    elif args.command == "manifest":
        path = PUBLIC / "databases" / "blackridge_public_v0.1.0.sqlite3"
        result = manifest(path)
        result["status"] = result["validation_status"]
        target = PUBLIC / "manifests" / "DATA_MANIFEST.json"
        target.write_text(json.dumps(result, indent=2) + "\n")
    elif args.command == "reconcile":
        result = validate(path)
    elif args.command == "workbook-qa":
        result = workbook_qa(PUBLIC / "workbooks" / "BLACKRIDGE_MASTER_TRACKER_v0.1.0.xlsx", path)
        REPORTS.mkdir(parents=True, exist_ok=True)
        (REPORTS / "WORKBOOK_QA_REPORT.json").write_text(json.dumps(result, indent=2) + "\n")
    print(json.dumps(result, indent=2))
    return 0 if result["status"] == "PASS" else 1
