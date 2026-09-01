from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session

from sable_harbor.accounting.models import Base, JournalEntry, Worker
from sable_harbor.generation import generate_baseline
from sable_harbor.reporting import build_workbook


def test_baseline_generation_is_idempotent_and_workbook_reopens(tmp_path: Path) -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        first = generate_baseline(session)
        session.commit()
        entry_count = session.scalar(select(func.count(JournalEntry.id)))
        second = generate_baseline(session)
        session.commit()
        assert second == first
        assert session.scalar(select(func.count(JournalEntry.id))) == entry_count
        assert (
            session.scalar(select(func.count(Worker.id)).where(Worker.worker_type == "EMPLOYEE"))
            == 708
        )
        output = build_workbook(session, tmp_path / "model.xlsx")
    workbook = load_workbook(output, data_only=False, read_only=True)
    assert {
        "Read Me",
        "Trial Balance",
        "Income Statement",
        "Balance Sheet",
        "Cash Flow Bridge",
        "Assumptions",
    }.issubset(workbook.sheetnames)
    assert workbook["Balance Sheet"]["B8"].value == "=B2-B6"
