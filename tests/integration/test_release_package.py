import json
import sqlite3
from copy import deepcopy
from datetime import UTC, date, datetime
from pathlib import Path

import pytest
from jsonschema import ValidationError  # type: ignore[import-untyped]
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from sable_harbor.accounting.models import (
    Account,
    AccountingBook,
    Base,
    EpistemicState,
    FactState,
    FiscalPeriod,
    LegalEntity,
    Site,
)
from sable_harbor.accounting.seed import seed_smoke
from sable_harbor.exports.release import (
    PUBLIC_ALLOWLIST,
    package_public_demo,
    sha256,
    validate_public_manifest,
)
from sable_harbor.exports.safety import scan_generated_artifacts
from sable_harbor.generation import generate_standard
from sable_harbor.provenance.service import complete_generation_run, record_generation_run


def test_public_release_manifest_inventory_and_checksums(tmp_path: Path) -> None:
    database = tmp_path / "source.sqlite"
    engine = create_engine(f"sqlite:///{database}")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        run = record_generation_run(
            session, profile="standard", scenario_code="base", seed=20260831, git_commit="a" * 40
        )
        generate_standard(session)
        complete_generation_run(session, run)
        unrelated_entity = LegalEntity(
            id="unrelated-entity",
            code="UNRELATED",
            name="Unrelated synthetic test entity",
            fact_state=FactState.SYNTHETIC_INSTANCE,
            existence_state=EpistemicState.SCENARIO,
            identity_state=EpistemicState.SCENARIO,
            relationship_state=EpistemicState.SCENARIO,
            effective_date_state=EpistemicState.SCENARIO,
            effective_from=date(2026, 1, 1),
            source_reference="test-only unrelated master",
            jurisdiction="OPEN",
        )
        unrelated_book = AccountingBook(
            id="unrelated-book",
            entity_id=unrelated_entity.id,
            code="PRIMARY_USD",
            currency="USD",
        )
        session.add_all(
            [
                unrelated_entity,
                unrelated_book,
                FiscalPeriod(
                    id="unrelated-period",
                    book_id=unrelated_book.id,
                    code="2099-01",
                    starts_on=date(2099, 1, 1),
                    ends_on=date(2099, 1, 31),
                ),
                Account(
                    id="unrelated-account",
                    code="9999",
                    name="Unrelated account",
                    account_class="ASSET",
                    normal_balance="DEBIT",
                ),
                Site(
                    id="unrelated-site",
                    code="UNRELATED-SITE",
                    name="Unrelated synthetic test site",
                    site_type="TEST",
                    region="OPEN",
                    owner_entity_id=unrelated_entity.id,
                    fact_state=FactState.SYNTHETIC_INSTANCE,
                ),
            ]
        )
        session.commit()
        generated_at = datetime(2026, 9, 5, 12, tzinfo=UTC)
        manifest_path = package_public_demo(
            session,
            tmp_path / "release",
            generation_run_id=run.id,
            generated_at=generated_at,
        )
    manifest = json.loads(manifest_path.read_text())
    validate_public_manifest(manifest)
    assert manifest["validation_status"] == "PASS"
    assert manifest["artifact_safety_scan"]["status"] == "PASS"
    assert manifest["artifact_safety_scan"]["failures"] == 0
    assert "PII_SHAPES" in manifest["artifact_safety_scan"]["scope"]
    assert manifest["schema_versions"] == ["0015"]
    assert manifest["classification"] == "PUBLIC_SAFE_SYNTHETIC"
    assert manifest["epistemic_mode"] == "RETROSPECTIVE_CURRENT_CANON"
    assert manifest["canon_effective_through"] == "2026-09-03"
    assert manifest["canon_reconciled_at"] == "2026-09-05"
    assert manifest["prepared_at"] == "2026-09-05"
    assert manifest["built_at"] == "2026-09-05T12:00:00+00:00"
    assert manifest["synthetic_calibration_through"] == "2026-08-31"
    assert "knowledge_cutoff" not in manifest
    assert manifest["input_version"] == "finance-generation-input-manifest/v1"
    assert len(manifest["input_manifest_digest"]) == 64
    assert len(manifest["generator_source_digest"]) == 64
    assert manifest["source_snapshot_ids"]["current_canon"].startswith("main@")
    assert (
        manifest["source_snapshot_digests"]["current_canon_git_commit"]
        == "712076751a31534cd9e6e41458336cdc7b6585b5"
    )
    assert {item["data_role"] for item in manifest["included_runs"]} == {
        "selected_synthetic_scenario",
        "shared_synthetic_calibration",
    }
    assert all(item["profile"] != "actual_common" for item in manifest["included_runs"])
    assert manifest["row_counts"]["journal_entry"] > 0
    assert manifest["output_hashes"] == {
        artifact["path"]: artifact["sha256"] for artifact in manifest["artifacts"]
    }
    malformed = deepcopy(manifest)
    malformed["built_at"] = "not-a-date"
    with pytest.raises(ValidationError):
        validate_public_manifest(malformed)
    malformed = deepcopy(manifest)
    malformed["forecast_from"] = "2026-99-99"
    with pytest.raises(ValidationError):
        validate_public_manifest(malformed)
    malformed = deepcopy(manifest)
    malformed["source_snapshot_ids"]["uncontrolled"] = "surprise"
    with pytest.raises(ValidationError):
        validate_public_manifest(malformed)
    malformed = deepcopy(manifest)
    malformed["row_counts"]["journal_entry"] = -1
    with pytest.raises(ValidationError):
        validate_public_manifest(malformed)
    malformed = deepcopy(manifest)
    malformed["artifacts"][0]["path"] = "../escape.csv"
    with pytest.raises(ValidationError):
        validate_public_manifest(malformed)
    for artifact in manifest["artifacts"]:
        path = manifest_path.parent / artifact["path"]
        assert path.exists()
        assert sha256(path) == artifact["sha256"]
    assert (manifest_path.parent / "sable_harbor_public_demo.sqlite").stat().st_size > 0
    allowlist = json.loads(PUBLIC_ALLOWLIST.read_text())["tables"]
    connection = sqlite3.connect(manifest_path.parent / "sable_harbor_public_demo.sqlite")
    try:
        actual_tables = {
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
            )
        }
        assert actual_tables == set(allowlist)
        for table, expected_columns in allowlist.items():
            actual_columns = [row[1] for row in connection.execute(f'PRAGMA table_info("{table}")')]
            assert actual_columns == expected_columns
        assert "generation_run" not in actual_tables
        assert "artifact" not in actual_tables
        for table, identifier in (
            ("legal_entity", "unrelated-entity"),
            ("accounting_book", "unrelated-book"),
            ("fiscal_period", "unrelated-period"),
            ("account", "unrelated-account"),
            ("site", "unrelated-site"),
        ):
            assert connection.execute(
                f'SELECT COUNT(*) FROM "{table}" WHERE id = ?', (identifier,)
            ).fetchone() == (0,)
    finally:
        connection.close()
    workbook = load_workbook(
        manifest_path.parent / "workbooks/SABLE_HARBOR_CONSOLIDATED_OPERATING_MODEL_v0.1.xlsx",
        read_only=True,
        data_only=False,
    )
    checks = workbook["Checks"]
    scoped_account_count = next(
        row[1].value
        for row in checks.iter_rows(min_row=9, max_col=2)
        if row[0].value == "Accounts loaded"
    )
    assert scoped_account_count == manifest["row_counts"]["account"]
    assert scan_generated_artifacts(manifest_path.parent) == []
    checksum_lines = (manifest_path.parent / "SHA256SUMS.txt").read_text().splitlines()
    assert any(line.endswith("  manifest.json") for line in checksum_lines)
    for line in checksum_lines:
        expected, relative = line.split("  ", 1)
        assert sha256(manifest_path.parent / relative) == expected


def test_public_release_is_deterministic_and_removes_stale_outputs(tmp_path: Path) -> None:
    database = tmp_path / "source.sqlite"
    engine = create_engine(f"sqlite:///{database}")
    Base.metadata.create_all(engine)
    generated_at = datetime(2026, 9, 5, 12, tzinfo=UTC)
    with Session(engine) as session:
        run = record_generation_run(
            session, profile="standard", scenario_code="base", seed=20260831, git_commit="a" * 40
        )
        generate_standard(session)
        complete_generation_run(session, run)
        session.commit()
        first = package_public_demo(
            session, tmp_path / "first", generation_run_id=run.id, generated_at=generated_at
        )
        second = package_public_demo(
            session, tmp_path / "second", generation_run_id=run.id, generated_at=generated_at
        )
        stale = tmp_path / "second" / "obsolete.txt"
        stale.write_text("must be removed")
        second = package_public_demo(
            session, tmp_path / "second", generation_run_id=run.id, generated_at=generated_at
        )
    first_files = {
        path.relative_to(first.parent): sha256(path)
        for path in first.parent.rglob("*")
        if path.is_file()
    }
    second_files = {
        path.relative_to(second.parent): sha256(path)
        for path in second.parent.rglob("*")
        if path.is_file()
    }
    assert first_files == second_files
    assert not stale.exists()


def test_public_release_rejects_unowned_or_broad_destination(tmp_path: Path) -> None:
    unowned = tmp_path / "unowned"
    unowned.mkdir()
    (unowned / "user-data.txt").write_text("preserve me")
    with pytest.raises(ValueError, match="unowned external directory"):
        package_public_demo(None, unowned, generation_run_id="unused")  # type: ignore[arg-type]
    assert (unowned / "user-data.txt").read_text() == "preserve me"

    with pytest.raises(ValueError, match="protected package destination"):
        package_public_demo(None, Path.cwd(), generation_run_id="unused")  # type: ignore[arg-type]


def test_public_release_rejects_a_completed_nonstandard_profile(tmp_path: Path) -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        run = record_generation_run(
            session, profile="smoke", scenario_code="base", seed=7, git_commit="a" * 40
        )
        seed_smoke(session, complete=False)
        complete_generation_run(session, run)
        session.commit()
        with pytest.raises(ValueError, match="standard profile"):
            package_public_demo(
                session,
                tmp_path / "invalid-release",
                generation_run_id=run.id,
                generated_at=run.completed_at,
            )


def test_public_release_rejects_a_stale_completed_build(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import sable_harbor.exports.metadata as metadata

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        run = record_generation_run(
            session, profile="standard", scenario_code="base", seed=7, git_commit="a" * 40
        )
        generate_standard(session, seed=7)
        complete_generation_run(session, run)
        session.commit()
        monkeypatch.setattr(metadata, "generation_input_manifest_digest", lambda: "f" * 64)
        with pytest.raises(ValueError, match="does not match the current governed build inputs"):
            package_public_demo(
                session,
                tmp_path / "stale-release",
                generation_run_id=run.id,
                generated_at=run.completed_at,
            )
    assert not (tmp_path / "stale-release").exists()


def test_generated_artifact_scan_reads_values_and_xlsx_relationships(tmp_path: Path) -> None:
    bad_csv = tmp_path / "bad.csv"
    bad_csv.write_text("value\n/home/private/oracle\n")
    failures = scan_generated_artifacts(tmp_path)
    assert any("forbidden marker" in failure for failure in failures)

    formula_csv = tmp_path / "formula.csv"
    formula_csv.write_text('value\n=WEBSERVICE("https://example.invalid")\n')
    assert any("formula injection" in failure for failure in scan_generated_artifacts(formula_csv))

    formula_csv.write_text('value\n\'=WEBSERVICE("https://example.invalid")\n')
    assert scan_generated_artifacts(formula_csv) == []


def test_generated_artifact_scan_reads_sqlite_text_without_raw_page_false_positives(
    tmp_path: Path,
) -> None:
    clean_database = tmp_path / "clean.sqlite3"
    with sqlite3.connect(clean_database) as connection:
        connection.execute("CREATE TABLE evidence (value TEXT NOT NULL)")
        connection.execute("INSERT INTO evidence VALUES (?)", ("synthetic public evidence",))
    assert scan_generated_artifacts(clean_database) == []

    with sqlite3.connect(clean_database) as connection:
        connection.execute("INSERT INTO evidence VALUES (?)", ("person@example.invalid",))
    assert any(
        "email-address-shaped value" in failure
        for failure in scan_generated_artifacts(clean_database)
    )


@pytest.mark.parametrize(
    "tables",
    (
        {"../escape": ["id"]},
        {"legal_entity": ['id" FROM generation_run --']},
    ),
)
def test_public_allowlist_rejects_unsafe_identifiers(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, tables: dict[str, list[str]]
) -> None:
    import sable_harbor.exports.release as release

    allowlist = tmp_path / "allowlist.json"
    allowlist.write_text(json.dumps({"version": "0.1.0", "tables": tables}))
    monkeypatch.setattr(release, "PUBLIC_ALLOWLIST", allowlist)
    with pytest.raises(ValueError, match="unsafe"):
        release._allowlist()


def test_staged_package_publish_rolls_back_and_rejects_symlinks(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import sable_harbor.exports.safety as safety

    destination = tmp_path / "owned-package"
    with safety.staged_package_directory(
        destination,
        package_kind="test-package",
        repository_output_root=tmp_path / "governed",
    ) as (_final, staging):
        (staging / "evidence.txt").write_text("old")

    real_replace = safety.os.replace

    def fail_staging_publish(source: object, target: object) -> None:
        source_path = Path(source)  # type: ignore[arg-type]
        target_path = Path(target)  # type: ignore[arg-type]
        if ".staging-" in source_path.name and target_path == destination:
            raise OSError("injected publish failure")
        real_replace(source_path, target_path)

    monkeypatch.setattr(safety.os, "replace", fail_staging_publish)
    with pytest.raises(OSError, match="injected publish failure"):
        with safety.staged_package_directory(
            destination,
            package_kind="test-package",
            repository_output_root=tmp_path / "governed",
        ) as (_final, staging):
            (staging / "evidence.txt").write_text("new")
    assert (destination / "evidence.txt").read_text() == "old"

    real_directory = tmp_path / "real-directory"
    real_directory.mkdir()
    symlink = tmp_path / "package-symlink"
    symlink.symlink_to(real_directory, target_is_directory=True)
    with pytest.raises(ValueError, match="symbolic link"):
        with safety.staged_package_directory(
            symlink,
            package_kind="test-package",
            repository_output_root=tmp_path / "governed",
        ):
            pass
