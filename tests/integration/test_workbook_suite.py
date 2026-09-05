import hashlib
import json
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from sable_harbor.accounting.models import Base
from sable_harbor.generation import generate_standard
from sable_harbor.provenance.service import complete_generation_run, record_generation_run
from sable_harbor.reporting_queries import run_named_query
from sable_harbor.workbooks.suite import (
    SHEET_SPECS,
    WORKBOOKS,
    generate_workbook_suite,
)


def test_six_workbooks_reopen_have_required_sheets_and_clean_links(tmp_path: Path) -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        run = record_generation_run(
            session, profile="standard", scenario_code="base", seed=20260831, git_commit="a" * 40
        )
        generate_standard(session)
        complete_generation_run(session, run)
        session.commit()
        expected_rows = {
            sheet: run_named_query(session, SHEET_SPECS[sheet].query, run.id)
            for sheet in (
                "Monthly Consolidated P&L",
                "Journal Detail Extract",
            )
        }
        expected_headers = {sheet: list(rows[0]) for sheet, rows in expected_rows.items()}
        expected_headers["Monthly Balance Sheet"] = [
            "period",
            "assets",
            "liabilities",
            "equity",
            "balance_sheet_difference",
        ]
        outputs = generate_workbook_suite(
            session,
            tmp_path,
            generation_run_id=run.id,
            generated_at=datetime(2026, 9, 5, 12, tzinfo=UTC),
        )
    manifest = json.loads((tmp_path / "workbook-suite-manifest.json").read_text())
    assert manifest["epistemic_mode"] == "RETROSPECTIVE_CURRENT_CANON"
    assert manifest["canon_effective_through"] == "2026-09-03"
    assert manifest["canon_reconciled_at"] == "2026-09-05"
    assert manifest["prepared_at"] == "2026-09-05"
    assert manifest["built_at"] == "2026-09-05T12:00:00+00:00"
    assert manifest["synthetic_calibration_through"] == "2026-08-31"
    assert "knowledge_cutoff" not in manifest
    assert manifest["input_version"] == "finance-generation-input-manifest/v1"
    assert len(manifest["input_manifest_digest"]) == 64
    assert manifest["artifact_safety_scan"]["status"] == "PASS"
    assert manifest["artifact_safety_scan"]["failures"] == 0
    assert manifest["output_hashes"] == {
        item["path"]: item["sha256"] for item in manifest["workbooks"]
    }
    for item in manifest["workbooks"]:
        assert hashlib.sha256((tmp_path / item["path"]).read_bytes()).hexdigest() == item["sha256"]
    checksum_lines = (tmp_path / "SHA256SUMS.txt").read_text().splitlines()
    assert any(line.endswith("  workbook-suite-manifest.json") for line in checksum_lines)
    for line in checksum_lines:
        expected, relative = line.split("  ", 1)
        assert hashlib.sha256((tmp_path / relative).read_bytes()).hexdigest() == expected
    assert set(SHEET_SPECS) == {sheet for sheets in WORKBOOKS.values() for sheet in sheets}
    assert {path.name for path in outputs} == set(WORKBOOKS)
    for output in outputs:
        assert 5_000 < output.stat().st_size < 20_000_000
        with ZipFile(output) as archive:
            assert not any("externalLinks" in name for name in archive.namelist())
        workbook = load_workbook(output, read_only=False, data_only=False)
        assert workbook.sheetnames == WORKBOOKS[output.name]
        check_sheet = workbook["Checks"]
        formulas = [
            cell.value for row in check_sheet.iter_rows() for cell in row if cell.data_type == "f"
        ]
        assert formulas
        assert all("#REF!" not in formula for formula in formulas)
        assert any("ABS(" in formula and "<=0.01" in formula for formula in formulas)
        for sheet_name, headers in expected_headers.items():
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            actual_headers = [cell.value for cell in worksheet[8]]
            assert actual_headers[: len(headers)] == headers
            if sheet_name not in expected_rows:
                continue
            assert worksheet.max_row == 8 + len(expected_rows[sheet_name])
            expected_first_row = [expected_rows[sheet_name][0][header] for header in headers]
            expected_first_row = [
                (
                    float(value)
                    if isinstance(value, Decimal)
                    else value.isoformat()
                    if isinstance(value, (date, datetime))
                    else value
                )
                for value in expected_first_row
            ]
            actual_first_row = [cell.value for cell in worksheet[9]][: len(headers)]
            assert actual_first_row == expected_first_row
        run_control = {
            row[0].value: row[1].value
            for row in workbook["Run Control"].iter_rows(min_row=9, max_col=2)
            if row[0].value is not None
        }
        assert run_control["epistemic_mode"] == "RETROSPECTIVE_CURRENT_CANON"
        assert run_control["synthetic_calibration_through"] == "2026-08-31"
        assert run_control["canon_effective_through"] == "2026-09-03"
        assert run_control["canon_reconciled_at"] == "2026-09-05"
        assert "knowledge_cutoff" not in run_control
        if "Generation Runs" in workbook.sheetnames:
            generation_runs_sheet = workbook["Generation Runs"]
            generation_run_headers = [cell.value for cell in generation_runs_sheet[8]]
            assert "synthetic_calibration_through" in generation_run_headers
            assert "actual_through" not in generation_run_headers
            profile_column = generation_run_headers.index("profile") + 1
            role_column = generation_run_headers.index("data_role") + 1
            profiles = {
                row[0].value
                for row in generation_runs_sheet.iter_rows(
                    min_row=9, min_col=profile_column, max_col=profile_column
                )
                if row[0].value is not None
            }
            roles = {
                row[0].value
                for row in generation_runs_sheet.iter_rows(
                    min_row=9, min_col=role_column, max_col=role_column
                )
                if row[0].value is not None
            }
            assert "shared_synthetic_calibration" in profiles
            assert "shared_synthetic_calibration" in roles
            assert profiles.isdisjoint({"actual_common", "synthetic_common"})


def test_sheet_routing_uses_exact_registry_not_title_substrings() -> None:
    assert SHEET_SPECS["Monthly Balance Sheet"].query == "@monthly_balance_sheet"
    assert SHEET_SPECS["Monthly Consolidated P&L"].query == "consolidated_monthly_pnl"
    assert SHEET_SPECS["Journal Detail Extract"].query == "journal_to_source_trace"
    assert SHEET_SPECS["Red Wash Unit Cost"].query == "red_wash_unit_cost_bridge"
    assert SHEET_SPECS["BS&T Route Customer Margin"].query == "bst_route_customer_margin"
    assert all(spec.purpose and spec.empty_state for spec in SHEET_SPECS.values())


def test_workbook_writer_rejects_rows_beyond_excel_limit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import sable_harbor.workbooks.suite as suite

    monkeypatch.setattr(suite, "EXCEL_MAX_ROWS", 10)
    with pytest.raises(ValueError, match="exceeds the Excel row limit"):
        suite._write_rows(
            None,
            ({"value": value} for value in (1, 2, 3)),
            None,
            None,
            "unused",
        )
